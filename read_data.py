import openpyxl

sheet = 'myexel.xlsx'

loaded_workbook = openpyxl.load_workbook(sheet)
# data_only used when there is a result on one column by computation of other columns
#loaded_workbook = openpyxl.load_workbook(sheet,data_only=True)
print(loaded_workbook)

#printing all sheets in the excel file
print(loaded_workbook.sheetnames)

for i in loaded_workbook:
    print(i)   #printing worksheet title 
    print(i.title)   # printing worksheet title in string format
    

s1 = 'Sheet1'
s2 = 'Sheet2'

# assigning Sheet1 to x
load_sheet_1 = loaded_workbook['Sheet1']
load_sheet_2 = loaded_workbook[s2]
print(f'this sheet is {load_sheet_1}')
# now all future programing will work on this sheet only
# x = a.active
 
b2_cell = load_sheet_1['B2']
c2_cell = load_sheet_1['C2']

print('-----------------')
print(b2_cell.value)
print(b2_cell)
print(c2_cell.value)

#y = a.active
b2_cell2 = load_sheet_2['B2']
c2_cell2 = load_sheet_2['C2']

print(b2_cell2, c2_cell2)
print(b2_cell2.value,c2_cell2.value)



print('++++++++++++')

#find value using coumn cordinates
print(load_sheet_1.cell(row=2,column=2).value)
#find find cordinates using values
print(c2_cell.row,c2_cell.column)


print(load_sheet_1['A2'].data_type)
print(load_sheet_1['B2'].data_type)

# n - for numeric 
# s - for string

# print the type of encoding
print(load_sheet_1['B2'].encoding)

print(load_sheet_1['b2'].parent)


cell_range = load_sheet_1['B2:C5']

cell_range = load_sheet_1['A1:C2']

for name,phone,email in cell_range:
    print(f'name: {name.value} , phone: {phone.value} ,email:{email.value}')
    
print('--------------------')
 
#finds which dimentions did the data occupied
print(f'Sheet Dimentions: {load_sheet_1.dimensions}')

# printing num of rows and column
print(load_sheet_1.max_row,load_sheet_1.max_column)

#print each column
for a,b,c,d,e,f in load_sheet_1[load_sheet_1.dimensions]:
    print(a.value,b.value,c.value,d.value,e.value,f.value)

print('++++++++++++++')

for row in load_sheet_1:
    for cell in row:
        print(f'{cell.value} -',end='')
    print('\n')
    
    
print('||||||||||||')

for row in load_sheet_1.rows:
    for cell in row:
        print(f'{cell.value} ',end='')
    print('\n')

print('?????????????\n\n')


for i in load_sheet_1.values:
    print(i)









